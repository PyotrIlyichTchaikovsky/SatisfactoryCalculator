#!/usr/bin/env python3
"""Copy FModel-exported planner icons into recipe_web/data/icons."""

from __future__ import annotations

import argparse
import json
import shutil
import struct
import sys
import zlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

import openpyxl

import game_rawdata_exporter as raw


SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_FILE_NAME = "raw_to_data.config.json"


class IconCopyError(RuntimeError):
    pass


@dataclass(frozen=True)
class DeviceIconRow:
    device_class: str
    icon_asset: str
    icon_path: str


@dataclass(frozen=True)
class MaterialIconRow:
    material_class: str
    material_category: str
    icon_path: str


def main(argv: Sequence[str] | None = None) -> int:
    try:
        args = parse_args(argv)
        config_path = args.config.expanduser().resolve()
        config = load_config(config_path)
        copy_devices = icon_copy_enabled(config, "copy_device_icons") or args.force
        copy_materials = icon_copy_enabled(config, "copy_material_icons") or args.force
        if not copy_devices and not copy_materials:
            print("Icon copy skipped: icons.copy_device_icons and icons.copy_material_icons are false.")
            return 0

        data_xlsx = args.data_xlsx or required_config_path(config, config_path, "output", "data_xlsx")
        exports_dir = args.exports_dir or optional_config_path(config, config_path, "icons", "fmodel_exports_dir")
        if exports_dir is None:
            raise IconCopyError("icons.fmodel_exports_dir is not configured")
        raw_json = required_config_path(config, config_path, "source", "raw_json") if copy_materials else None

        missing_count = 0
        if copy_devices:
            result = copy_device_icons(data_xlsx.resolve(), exports_dir.resolve(), dry_run=args.dry_run)
            print(
                "Device icons: "
                f"copied={result['copied']}, unchanged={result['unchanged']}, missing={result['missing']}"
            )
            missing_count += result["missing"]
        else:
            result = {"missing_rows": []}

        material_result: dict[str, Any] = {"missing_rows": []}
        if copy_materials:
            material_result = copy_material_icons(
                data_xlsx.resolve(),
                exports_dir.resolve(),
                raw_json.resolve() if raw_json else None,
                dry_run=args.dry_run,
            )
            print(
                "Material icons: "
                f"copied={material_result['copied']}, unchanged={material_result['unchanged']}, "
                f"generated={material_result['generated']}, missing={material_result['missing']}"
            )
            missing_count += material_result["missing"]

        if result.get("missing_rows"):
            print("Missing device icons:", file=sys.stderr)
            for row in result["missing_rows"]:
                print(f"  {row.device_class}: {row.icon_asset}", file=sys.stderr)
        if material_result.get("missing_rows"):
            print("Missing material icons:", file=sys.stderr)
            for row in material_result["missing_rows"]:
                print(f"  {row.material_class}: {row.icon_path}", file=sys.stderr)
        if missing_count:
            return 1
        return 0
    except (IconCopyError, OSError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2


def parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Copy FModel-exported device icons into Data.xlsx icon paths.")
    parser.add_argument(
        "--config",
        type=Path,
        default=SCRIPT_DIR / CONFIG_FILE_NAME,
        help="Path to raw_to_data.config.json.",
    )
    parser.add_argument("--data-xlsx", type=Path, help="Override Data.xlsx path.")
    parser.add_argument("--exports-dir", type=Path, help="Override FModel Output/Exports directory.")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be copied without writing files.")
    parser.add_argument("--force", action="store_true", help="Run even when icons.copy_device_icons is false.")
    return parser.parse_args(argv)


def load_config(path: Path) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            config = json.load(handle)
    except FileNotFoundError as exc:
        raise IconCopyError(f"config file does not exist: {path}") from exc
    except json.JSONDecodeError as exc:
        raise IconCopyError(f"config file is not valid JSON: {path}: {exc}") from exc
    if not isinstance(config, dict):
        raise IconCopyError(f"config file must contain a JSON object: {path}")
    return config


def icon_copy_enabled(config: dict[str, Any], key: str) -> bool:
    section = config.get("icons", {})
    if section is None:
        return False
    if not isinstance(section, dict):
        raise IconCopyError("icons must be an object")
    return bool(section.get(key, False))


def required_config_path(config: dict[str, Any], config_path: Path, section: str, key: str) -> Path:
    path = optional_config_path(config, config_path, section, key)
    if path is None:
        raise IconCopyError(f"config key {section}.{key} must be a non-empty path string")
    return path


def optional_config_path(config: dict[str, Any], config_path: Path, section: str, key: str) -> Path | None:
    section_value = config.get(section)
    if not isinstance(section_value, dict):
        return None
    value = section_value.get(key)
    if value is None or value == "":
        return None
    if not isinstance(value, str):
        raise IconCopyError(f"config key {section}.{key} must be a path string")
    path = Path(value).expanduser()
    if path.is_absolute():
        return path
    return (config_path.parent / path).resolve()


def copy_device_icons(data_xlsx: Path, exports_dir: Path, dry_run: bool = False) -> dict[str, Any]:
    if not data_xlsx.exists():
        raise IconCopyError(f"Data.xlsx does not exist: {data_xlsx}")
    if not exports_dir.exists():
        raise IconCopyError(f"FModel exports directory does not exist: {exports_dir}")

    rows = load_device_icon_rows(data_xlsx)
    png_index = build_png_index(exports_dir)
    data_dir = data_xlsx.parent

    copied = 0
    unchanged = 0
    missing_rows: list[DeviceIconRow] = []

    for row in rows:
        source = find_exported_icon(row.icon_asset, exports_dir, png_index)
        if source is None:
            missing_rows.append(row)
            continue

        target = (data_dir / row.icon_path).resolve()
        if same_file_content(source, target):
            unchanged += 1
            continue

        if dry_run:
            print(f"copy {source} -> {target}")
        else:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
        copied += 1

    return {
        "copied": copied,
        "unchanged": unchanged,
        "missing": len(missing_rows),
        "missing_rows": missing_rows,
    }


def copy_material_icons(data_xlsx: Path, exports_dir: Path, raw_json: Path | None, dry_run: bool = False) -> dict[str, Any]:
    if not data_xlsx.exists():
        raise IconCopyError(f"Data.xlsx does not exist: {data_xlsx}")
    if raw_json is None or not raw_json.exists():
        raise IconCopyError(f"RawData.json does not exist: {raw_json}")
    if not exports_dir.exists():
        raise IconCopyError(f"FModel exports directory does not exist: {exports_dir}")

    rows = load_material_icon_rows(data_xlsx)
    material_icon_assets = load_material_icon_assets(raw_json)
    png_index = build_png_index(exports_dir)
    data_dir = data_xlsx.parent

    copied = 0
    unchanged = 0
    generated = 0
    missing_rows: list[MaterialIconRow] = []

    for row in rows:
        target = (data_dir / row.icon_path).resolve()
        if row.material_category == raw.MATERIAL_CATEGORY_POWER:
            if ensure_power_icon(target, dry_run=dry_run):
                generated += 1
            else:
                unchanged += 1
            continue

        icon_asset = material_icon_assets.get(row.material_class) or material_icon_assets.get(row.material_class.removesuffix("_C"))
        if not icon_asset:
            missing_rows.append(row)
            continue
        source = find_exported_icon(icon_asset, exports_dir, png_index)
        if source is None:
            missing_rows.append(row)
            continue

        if same_file_content(source, target):
            unchanged += 1
            continue

        if dry_run:
            print(f"copy {source} -> {target}")
        else:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
        copied += 1

    return {
        "copied": copied,
        "unchanged": unchanged,
        "generated": generated,
        "missing": len(missing_rows),
        "missing_rows": missing_rows,
    }


def load_device_icon_rows(data_xlsx: Path) -> list[DeviceIconRow]:
    workbook = openpyxl.load_workbook(data_xlsx, data_only=True, read_only=True)
    if "Devices" not in workbook.sheetnames:
        raise IconCopyError(f"Devices sheet not found: {data_xlsx}")
    sheet = workbook["Devices"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {str(header): idx for idx, header in enumerate(headers) if header}
    required = ("DeviceClassName", "IconAsset", "IconPath")
    missing = [name for name in required if name not in index]
    if missing:
        raise IconCopyError(f"Devices sheet is missing columns: {', '.join(missing)}")

    rows: list[DeviceIconRow] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        device_class = string_cell(values[index["DeviceClassName"]])
        icon_asset = string_cell(values[index["IconAsset"]])
        icon_path = string_cell(values[index["IconPath"]])
        if device_class and icon_asset and icon_path:
            rows.append(DeviceIconRow(device_class=device_class, icon_asset=icon_asset, icon_path=icon_path))
    return rows


def load_material_icon_rows(data_xlsx: Path) -> list[MaterialIconRow]:
    workbook = openpyxl.load_workbook(data_xlsx, data_only=True, read_only=True)
    if "Materials" not in workbook.sheetnames:
        raise IconCopyError(f"Materials sheet not found: {data_xlsx}")
    sheet = workbook["Materials"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {str(header): idx for idx, header in enumerate(headers) if header}
    required = ("MaterialClassName", "MaterialCategory", "Icon")
    missing = [name for name in required if name not in index]
    if missing:
        raise IconCopyError(f"Materials sheet is missing columns: {', '.join(missing)}")

    rows: list[MaterialIconRow] = []
    seen: set[str] = set()
    last_class = ""
    last_category = ""
    last_icon = ""
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row_class = string_cell(values[index["MaterialClassName"]])
        row_category = string_cell(values[index["MaterialCategory"]])
        row_icon = string_cell(values[index["Icon"]])
        if row_class:
            material_class = row_class
            material_category = row_category
            icon_path = row_icon
            last_class = row_class
            last_category = row_category
            last_icon = row_icon
        else:
            material_class = last_class
            material_category = row_category or last_category
            icon_path = row_icon or last_icon
        if not material_class or not icon_path or material_class in seen:
            continue
        seen.add(material_class)
        rows.append(
            MaterialIconRow(
                material_class=material_class,
                material_category=material_category,
                icon_path=icon_path,
            )
        )
    return rows


def load_material_icon_assets(raw_json: Path) -> dict[str, str]:
    data = raw.load_json(raw_json)
    classes = list(raw.iter_docs_classes(data))
    display_names = raw.build_display_name_index(classes)
    items = raw.build_item_index(classes, display_names)
    return {
        key: item.icon_asset
        for key, item in items.items()
        if item.icon_asset
    }


def string_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def build_png_index(exports_dir: Path) -> dict[str, list[Path]]:
    result: dict[str, list[Path]] = {}
    for path in exports_dir.rglob("*.png"):
        result.setdefault(path.name.lower(), []).append(path)
    for paths in result.values():
        paths.sort(key=lambda item: (len(str(item)), str(item).lower()))
    return result


def find_exported_icon(icon_asset: str, exports_dir: Path, png_index: dict[str, list[Path]]) -> Path | None:
    object_path = unreal_object_path(icon_asset)
    package_path = object_path.split(".", 1)[0]
    asset_name = package_path.rsplit("/", 1)[-1]

    for candidate in exact_fmodel_paths(exports_dir, package_path, asset_name):
        if candidate.exists():
            return candidate

    matches = png_index.get(f"{asset_name}.png".lower(), [])
    return matches[0] if matches else None


def unreal_object_path(icon_asset: str) -> str:
    value = icon_asset.strip()
    if " " in value:
        value = value.split(" ", 1)[1].strip()
    if not value.startswith("/"):
        raise IconCopyError(f"unsupported icon asset path: {icon_asset}")
    return value


def exact_fmodel_paths(exports_dir: Path, package_path: str, asset_name: str) -> list[Path]:
    if package_path.startswith("/Game/"):
        relative = Path("FactoryGame") / "Content" / package_path[len("/Game/") :]
    else:
        relative = Path(package_path.lstrip("/"))
    return [
        exports_dir / relative.parent / f"{asset_name}.png",
        exports_dir / relative / f"{asset_name}.png",
    ]


def same_file_content(source: Path, target: Path) -> bool:
    if not target.exists():
        return False
    source_stat = source.stat()
    target_stat = target.stat()
    if source_stat.st_size != target_stat.st_size:
        return False
    return source.read_bytes() == target.read_bytes()


def ensure_power_icon(target: Path, dry_run: bool = False) -> bool:
    png_data = power_icon_png()
    if target.exists() and target.read_bytes() == png_data:
        return False
    if dry_run:
        print(f"generate power icon -> {target}")
    else:
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(png_data)
    return True


def power_icon_png(size: int = 128) -> bytes:
    pixels: list[tuple[int, int, int, int]] = []
    center = (size - 1) / 2
    radius = size * 0.43
    inner_radius = size * 0.35
    bolt = [
        (size * 0.54, size * 0.10),
        (size * 0.30, size * 0.56),
        (size * 0.48, size * 0.56),
        (size * 0.39, size * 0.92),
        (size * 0.72, size * 0.43),
        (size * 0.53, size * 0.43),
    ]
    for y in range(size):
        for x in range(size):
            dx = x - center
            dy = y - center
            distance = (dx * dx + dy * dy) ** 0.5
            if point_in_polygon(x + 0.5, y + 0.5, bolt):
                pixels.append((255, 222, 73, 255))
            elif distance <= inner_radius:
                pixels.append((28, 83, 117, 255))
            elif distance <= radius:
                pixels.append((17, 44, 64, 255))
            else:
                pixels.append((0, 0, 0, 0))
    return rgba_png(size, size, pixels)


def point_in_polygon(x: float, y: float, polygon: Sequence[tuple[float, float]]) -> bool:
    inside = False
    j = len(polygon) - 1
    for i, point in enumerate(polygon):
        xi, yi = point
        xj, yj = polygon[j]
        if ((yi > y) != (yj > y)) and x < ((xj - xi) * (y - yi) / ((yj - yi) or 1e-9) + xi):
            inside = not inside
        j = i
    return inside


def rgba_png(width: int, height: int, pixels: Sequence[tuple[int, int, int, int]]) -> bytes:
    def chunk(kind: bytes, payload: bytes) -> bytes:
        return (
            struct.pack(">I", len(payload))
            + kind
            + payload
            + struct.pack(">I", zlib.crc32(kind + payload) & 0xFFFFFFFF)
        )

    rows = []
    for y in range(height):
        row = bytearray([0])
        for pixel in pixels[y * width : (y + 1) * width]:
            row.extend(pixel)
        rows.append(bytes(row))
    raw_data = b"".join(rows)
    header = struct.pack(">IIBBBBB", width, height, 8, 6, 0, 0, 0)
    return b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", header) + chunk(b"IDAT", zlib.compress(raw_data, 9)) + chunk(b"IEND", b"")


if __name__ == "__main__":
    raise SystemExit(main())
